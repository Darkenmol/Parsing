import csv
import openpyxl
import sys
from openpyxl import load_workbook

def xlsx_to_csv(input_file, output_file, delimiter=';'):
    # Загрузка xlsx файла
    wb = load_workbook(input_file)
    ws = wb.active
    


    # Открытие файла csv для записи
    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter=delimiter)

        # Обход строк и запись значений ячеек в csv
        for row in ws.iter_rows():                
            row_data = [cell.value.replace('\n','') if cell.value is not None else '' for cell in row]
            writer.writerow(row_data)

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Использование: python script.py input.xlsx") #  output.csv
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = input_file[:input_file.rfind('.')] + '.csv'
    # output_file = sys.argv[2]

    xlsx_to_csv(input_file, output_file)
